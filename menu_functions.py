from os import read
from matplotlib import pyplot as plt
from pandas.core.algorithms import mode
from typing import Tuple
import csv
from openpyxl import *
import pandas as pd
from colorama import init, Fore, Back, Style
from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import numpy as np


workbook=load_workbook(filename="Users.xlsx")
EmployeesSheet=workbook["Employees"]
ClientsSheet=workbook["Clients"]
VVX=load_workbook(filename="Tariffs.xlsx")
Tariffs=VVX["Tariffs"]

def Close(the_window):
    the_window.destroy()
    return

def showTariffs(window,count=0):
    tree=ttk.Treeview(window,columns=('ID','TN','Data','Time','Price'),show="headings",height=5)
    tree.column('#1',width=80)
    tree.column('#2',width=80)
    tree.column('#3',width=80)
    tree.column('#4',width=80)
    tree.column('#5',width=80)
    tree.heading('#1', text='ID')
    tree.heading('#2', text='Tariff Name')
    tree.heading('#3', text='Data')
    tree.heading('#4', text='Time')
    tree.heading('#5', text='Price')
    d=0
    tree.grid(columnspan=2,row=count)
    for yes in Tariffs.iter_rows(values_only=True,min_row=2):
        tree.insert('',d,values=yes)
        count+=1
    CloseButton=Button(window,text="Close",command=lambda w=window: Close(w)).grid(columnspan=2,row=count+1)

def showMyTariff(window,details):
    ActiveT=""
    PreviousT=""
    print(details)
    for col in Tariffs.iter_rows(min_col=1,max_col=3,min_row=2,values_only=True):
        if col[0]==details[5]:
            ActiveT=col[1]
    for col in Tariffs.iter_rows(min_col=1,max_col=3,min_row=2,values_only=True):
        if col[0]==details[6]:
            PreviousT=col[1]
    Lbl1=Label(window,text=f"Your active Tariff is {ActiveT}").grid(column=0,row=0,padx=80)
    Lbl2=Label(window,text=f"Your previous Tariff was {PreviousT}").grid(column=0,row=1)
    CloseButton=Button(window,text="Close",command=lambda w=window: Close(w)).grid(column=0,row=2)
    
def users_list(window,t):
    workbook = load_workbook(filename="Users.xlsx")
    EmployeesSheet = workbook["Employees"]
    ClientsSheet = workbook["Clients"]
    VVX = load_workbook(filename="Tariffs.xlsx")
    Tariffs = VVX["Tariffs"]
    tree=ttk.Treeview(window,columns=('ID','Name','Login','Fourth','F'),show="headings")
    tree.column('#1',width=75)
    tree.column('#2',width=75)
    tree.column('#3',width=75)
    tree.column('#4',width=75)
    tree.column('#5',width=0)
    tree.heading('#1', text='ID')
    tree.heading('#2', text='Name')
    tree.heading('#3', text='Login')
    tree.heading('#4', text='Balance')
    if t=='Client':
        type=ClientsSheet
    elif t=='Employee':
        type=EmployeesSheet
        tree.heading('#4', text='Position')
        tree.heading('#5', text='Salary')
        tree.column('#1',width=61)
        tree.column('#2',width=61)
        tree.column('#3',width=61)
        tree.column('#4',width=61)
        tree.column('#5',width=61)
    count=0
    tree.pack()
    for yes in type.iter_rows(values_only=True,min_row=2):
        if t=='Client':
            user=[yes[0],yes[1],yes[2],yes[4]]
        elif t=='Employee':
            user=[yes[0],yes[1],yes[2],yes[4],yes[5]]
        tree.insert('',count,values=user)
        count+=1
    CloseButton=Button(window,text="Close",command=lambda w=window: Close(w)).pack()
    
def searchClient(window,Search,tree):
    name=Search.get()
    found=False
    while True:
        for col in ClientsSheet.iter_rows(min_row=2,values_only=True):
            if (name.upper() in col[1].upper())==True:
                found=True
                fc=[col[1],col[2],col[3],col[4],col[7],col[8]]
                tree.insert('',1,values=fc)
        if found==False:
            messagebox.showinfo("Error", "Client could not be found")
            continue
        elif found==True:
            break

def searchClientWindow(window,Search):
    tree=ttk.Treeview(window,columns=('ID','Name','Login','Balance','Age','Balance'),show="headings")
    tree.column('#1',width=66)
    tree.column('#2',width=66)
    tree.column('#3',width=66)
    tree.column('#4',width=66)
    tree.column('#5',width=66)
    tree.column('#6',width=66)
    tree.heading('#1', text='ID')
    tree.heading('#2', text='Name')
    tree.heading('#3', text='Login')
    tree.heading('#4', text='Balance')
    tree.heading('#5', text='Age')
    tree.heading('#6', text='Balance')
    tree.grid(columnspan=2,row=2)
    SearchLabel=Label(window,text="Enter the name of a Client").grid(column=0,row=0)
    SearchField=Entry(window,textvariable=Search).grid(column=1,row=0)
    SearchButton=Button(window,text="Enter",command=lambda w=window,s=Search: searchClient(w,s,tree)).grid(column=0,row=1)
    CloseButton=Button(window,text="Close",command=lambda w=window: Close(w)).grid(column=1,row=1)
    
def historyUser(window,Search):
    ID=Search.get()
    for col in ClientsSheet.iter_rows(min_row=2,values_only=True):
        if ID==int(col[0]):
            Lbl1=Label(window,text=f"{col[1]}'s active Tariff is {col[5]}").grid(column=0,row=2)
            Lbl2=Label(window,text=f"{col[1]}'s previous Tariff was {col[6]}").grid(column=0,row=3)

def historyUserWindow(window,Search):
    SearchLabel=Label(window,text="Enter the ID of a Client").grid(column=0,row=0)
    SearchField=Entry(window,textvariable=Search).grid(column=1,row=0)
    SearchButton=Button(window,text="Enter",command=lambda w=window,s=Search: historyUser(w,s)).grid(columnspan=2,row=1)
    CloseButton=Button(window,text="Close",command=lambda w=window: Close(w)).grid(columnspan=2,row=4)
    
def sortClients(window):
    tree=ttk.Treeview(window,columns=('ID','Name','Age'),show="headings")
    tree.column('#1',width=133)
    tree.column('#2',width=133)
    tree.column('#3',width=133)
    tree.heading('#1', text='ID')
    tree.heading('#2', text='Name')
    tree.heading('#3', text='Age')
    MainLabel=Label(window,text="Sorting window").grid(columnspan=4,row=0)
    IDButton=Button(window,text="ID",command=lambda:IDSort(window,tree)).grid(column=0,row=1)
    NameButton=Button(window,text="Name",command=lambda:NameSort(window,tree)).grid(column=1,row=1)
    AgeButton=Button(window,text="Age",command=lambda:AgeSort(window,tree)).grid(column=2,row=1)
    tree.grid(columnspan=3,row=2)
    CloseButton=Button(window,text="Close",command=lambda w=window: Close(w)).grid(columnspan=3,row=3)
    
def AgeFun():
    gp = pd.read_excel('Users.xlsx',sheet_name="Clients")
    old = 0
    jun = 0
    med = 0
    for i in range(len(gp['Age'])):
        if gp['Age'][i] < 18:
            jun += 1
        elif 18 <= gp['Age'][i] < 50:
            med += 1
        else:
            old += 1
    data = [
        ['Young', jun],
        ['Mature', med],
        ['Old', old],
    ]
    values = [x[1] for x in data]
    labels = [x[0] for x in data]
    fig, ax = plt.subplots()
    fig.canvas.set_window_title('INAI')
    ax.pie(values, labels=labels, autopct="%.1f%%", radius=1.2)
    ax.set_aspect("equal")
    plt.show()

def stats(window):
    RegionButton=Button(window,text="Region Chart",command=lambda: graf()).grid(column=0,row=0,padx=150)
    AgeButton=Button(window,text="Age Pie", command=lambda: AgeFun()).grid(column=0,row=1)
    CloseButton=Button(window,text="Close",command=lambda w=window: Close(w)).grid(column=0,row=2)
    
def addBalance(Search,Amount,w):
    workbook = load_workbook(filename="Users.xlsx")
    ClientsSheet = workbook["Clients"]
    s=Search.get()
    a=Amount.get()
    count=1
    for col in ClientsSheet.iter_rows(min_row=2,values_only=True):
        if s==col[0]:
            ClientsSheet[f"E{count}"]=int(col[4])+a
            messagebox.showinfo("Balance","Money has been deposited")
            workbook.save("Users.xlsx")
        count+=1
    
def addBalanceWindow(window,Search,Amount):
    SearchLabel=Label(window,text="Enter the ID of a Client").grid(column=0,row=0)
    SearchField=Entry(window,textvariable=Search).grid(column=1,row=0)
    AmountLabel=Label(window,text="Enter the Amount").grid(column=0,row=1)
    AmountField=Entry(window,textvariable=Amount).grid(column=1,row=1)
    SearchButton=Button(window,text="Enter",command=lambda s=Search,a=Amount,w=window: addBalance(s,a,w)).grid(columnspan=2,row=2)
    CloseButton=Button(window,text="Close",command=lambda w=window: Close(w)).grid(columnspan=2,row=3)
    
def subscribeToNewTariff(window,details,idTariff):
    thefile=pd.DataFrame([details[0],idTariff])
    thefile.to_csv("Applications.csv",mode='a',header=False)
    Response=Label(window,text="Your request has been submitted").grid(column=0,row=2)

def subscribeToNewTariffWindow(window,idTariff,details):
    TariffLabel=Label(window,text="Enter the ID of a Tariff:").grid(column=0,row=0,padx=50)
    TariffField=Entry(window,textvariable=idTariff).grid(column=1,row=0)
    TariffButton=Button(window,text="Enter", command=lambda w=window, d=details,id=idTariff: subscribeToNewTariff(w,d,id)).grid(columnspan=2,row=1)
    showTariffs(window,count=2)

def RequestSubFunction1(id):
    for col in ClientsSheet.iter_rows(min_row=2,values_only=True):
        if id==col[0]:
            return col[1]

def RequestSubFunction2(id):
    for row in Tariffs.iter_rows(min_row=2,values_only=True):
        if row[0]==id:
            return row[1]

def viewListOfReqest(window):
    thefile=pd.read_csv("Applications.csv")
    count=0
    for row in range(len(thefile['clientId'])):
        q=[(thefile['clientId'][row]),(thefile['tariffId'][row])]
        infoClient=RequestSubFunction1(q[0])
        infoTariff=RequestSubFunction2(q[1])
        RequestLabel=Label(window,text=f"User {infoClient} wishes to use Tariff {infoTariff}").grid(column=0,row=count)
        ApproveButton=Button(window,text="Approve",command=lambda t="A",qq=q,c=count: analysisRequest(t,qq,c)).grid(column=1,row=count)
        RejectButton=Button(window,text="Reject",command=lambda t="R",qq=q,c=count: analysisRequest(t,qq,c)).grid(column=2,row=count)
        count+=1
    CloseButton=Button(window,text="Close",command=lambda w=window: Close(w)).grid(columnspan=3,row=count+1)
    
def analysisRequest(t,qq,c):
    workbook=load_workbook(filename="Users.xlsx")
    ClientsSheet=workbook["Clients"]
    thefile=pd.read_csv("Applications.csv")
    if t=="A":
        count=1
        for col in ClientsSheet.iter_rows(min_row=2,values_only=True):
            count+=1
            if qq[0]==col[0]:
                ClientsSheet[f"G{count}"]=col[5]
                ClientsSheet[f"F{count}"]=qq[1]
                workbook.save("Users.xlsx")
    messagebox.showinfo("Request","Request has been processed")
    thefile.drop(c)
    thefile.to_csv("Applications.csv")

def addInfoToClient(Age,Region,details):
    count=2
    print(details)
    workbook = load_workbook(filename="Users.xlsx")
    ClientsSheet = workbook["Clients"]
    a=Age.get()
    r=Region.get()
    if a!=0 or r!="":
        for col in ClientsSheet.iter_rows(min_row=2,values_only=True):
            if details[0]==col[0]:
                ClientsSheet[f"H{count}"]=a
                ClientsSheet[f"I{count}"]=r
                workbook.save("Users.xlsx")
            count+=1
        messagebox.showinfo("Success", "Extra info has been logged")
    else:
        messagebox.showinfo("Error", "One of the fields is not filled in")

def addInfoToClientWindow(details,window):
    Age=IntVar(window)
    Region=StringVar(window)
    AgeField=Entry(window, textvariable=Age).grid(column=1,row=0)
    AgeLabel=Label(window,text='Age').grid(column=0,row=0,padx=50)
    RegionField=Entry(window, textvariable=Region).grid(column=1,row=1)
    RegionLabel=Label(window,text='Region').grid(column=0,row=1)
    Enter_Button = Button(window, text="Enter",command=lambda :addInfoToClient(Age,Region,details)).grid(columnspan=2,row=2)
    CloseButton=Button(window,text="Close",command=lambda w=window: Close(w)).grid(columnspan=2,row=3)

def graf():
    gs = pd.read_excel('Users.xlsx',sheet_name="Clients")
    n = [0, 0, 0, 0, 0, 0, 0, ]
    for i in range(len(gs['City'])):
        if gs['City'][i] == 'Osh':
            n[0] += 1
        elif gs['City'][i] == 'Batken':
            n[1] += 1
        elif gs['City'][i] == 'Jalal-Abad':
            n[2] += 1
        elif gs['City'][i] == 'Chuy':
            n[3] += 1
        elif gs['City'][i] == 'Issuk-Kul':
            n[4] += 1
        elif gs['City'][i] == 'Talas':
            n[5] += 1
        elif gs['City'][i] == 'Naryn':
            n[6] += 1
        index = np.arange(7)

    plt.title('Region Chart')
    plt.bar(index, n, error_kw={'ecolor': '0.1', 'capsize': 8}, alpha=0.9, label='Regions')
    plt.xticks(index, ['Osh', 'Batken', 'Jalal-Abad', 'Chuy', 'Issuk-Kul', 'Talas', 'Naryn'])
    plt.legend(loc=2)
    plt.show()

def createEmployee(n,l,p,pp):
    workbook=load_workbook(filename="Users.xlsx")
    EmployeesSheet=workbook["Employees"]
    account_type=EmployeesSheet
    new_account=[int(len(account_type["A"])),n.get(),l.get(),p.get(),pp.get(),0]
    account_type.append(new_account)
    workbook.save("Users.xlsx")
    messagebox.showinfo("User","Employee has been created")

def createEmployeeWindow(window,Login,Password,Name,Position):
    NameField=Entry(window, textvariable=Name).grid(column=1,row=0)
    NameLabel=Label(window,text='Name').grid(column=0,row=0,padx=50)
    LoginField=Entry(window, textvariable=Login).grid(column=1,row=1)
    LoginLabel=Label(window,text='Login').grid(column=0,row=1)
    PasswordField=Entry(window, textvariable=Password).grid(column=1,row=2)
    PasswordLabel=Label(window,text='Password').grid(column=0,row=2)
    PositionField=Entry(window, textvariable=Position).grid(column=1,row=3)
    PositionLabel=Label(window,text='Position').grid(column=0,row=3)
    Enter_Button = Button(window, text="Enter",command=lambda t="Clients", n=Name,l=Login,p=Password,pp=Position: createEmployee(n,l,p,pp)).grid(columnspan=2,row=4)
    CloseButton=Button(window,text="Close",command=lambda w=window: Close(w)).grid(columnspan=2,row=5)

def IDSort(window,tree):
    tree.delete(*tree.get_children())
    b = pd.read_excel('Users.xlsx',sheet_name="Clients")
    lst = []
    for i in range(len(b["Id"])):
        lst.append([b["Id"][i], b["Name"][i], b["Age"][i], b["City"][i]])
    lst = sorted(lst, key=lambda x: x[0])
    count=0
    for i in lst:
        tree.insert('',count,values=i)
        count+=1

def NameSort(window,tree):
    tree.delete(*tree.get_children())
    b = pd.read_excel('Users.xlsx',sheet_name="Clients")
    lst = []
    for i in range(len(b["Id"])):
        lst.append([b["Id"][i], b["Name"][i], b["Age"][i], b["City"][i]])
    lst = sorted(lst, key=lambda x: x[1])
    count=0
    for i in lst:
        tree.insert('',count,values=i)
        count+=1

def AgeSort(window,tree):
    tree.delete(*tree.get_children())
    b = pd.read_excel('Users.xlsx',sheet_name="Clients")
    lst = []
    for i in range(len(b["Id"])):
        lst.append([b["Id"][i], b["Name"][i], b["Age"][i], b["City"][i]])
    lst = sorted(lst, key=lambda x: x[2])
    count=0
    for i in lst:
        tree.insert('',count,values=i)
        count+=1

def changeSalary(window,ss,id):
    workbook = load_workbook(filename="Users.xlsx")
    EmployeesSheet = workbook["Employees"]
    ID=id.get()
    Salary=ss.get()
    count=1
    for col in EmployeesSheet.iter_rows(min_row=2,values_only=True):
        if ID==col[0]:
            EmployeesSheet[f"F{count}"]=Salary
            messagebox.showinfo("Salary","Salary has been changed")
            workbook.save("Users.xlsx")
        count+=1
    
def changeSalaryWindow(window,Salary,ID):
    SearchLabel=Label(window,text="Enter the ID of an Employee").grid(column=0,row=0)
    SearchField=Entry(window,textvariable=ID).grid(column=1,row=0)
    AmountLabel=Label(window,text="Enter new Salary").grid(column=0,row=1)
    AmountField=Entry(window,textvariable=Salary).grid(column=1,row=1)
    SearchButton=Button(window,text="Enter",command=lambda i=ID,s=Salary,w=window: changeSalary(w,s,i)).grid(columnspan=2,row=2)
    CloseButton=Button(window,text="Close",command=lambda w=window: Close(w)).grid(columnspan=2,row=3)

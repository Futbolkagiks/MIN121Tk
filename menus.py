from typing import *
import csv
from numpy import log
from openpyxl import *
from colorama import init, Fore, Back, Style
import pandas as pd
from menu_functions import *
from tkinter import *
from tkinter import messagebox

def exit():
    quit()

def users_menu(col):
    workbook = load_workbook(filename="Users.xlsx")
    EmployeesSheet = workbook["Employees"]
    ClientsSheet = workbook["Clients"]
    VVX = load_workbook(filename="Tariffs.xlsx")
    Tariffs = VVX["Tariffs"]
    window = Tk()
    window.title("User's menu")
    window.geometry('305x250+550+200')
    window.resizable(width=False, height=False)
    window.configure(bg='#58D69B')

    def clicked1(col):
        window = Toplevel()
        window.title("MIN-1-21 Project")
        window.iconbitmap("image.ico")
        window.geometry('305x100+550+200')
        showMyTariff(window,col)

    def clicked2(col):
        window = Toplevel()
        window.title("MIN-1-21 Project")
        window.iconbitmap("image.ico")
        window.geometry('305x100+550+200')
        Balance=Label(window,text=f"Your balance is {col[4]}").pack(padx=50)
        CloseBtt=Button(window,text="Close",command=lambda w=window: Close(w)).pack()

    def clicked3(details):
        window = Toplevel()
        idTariff=IntVar(window)
        window.title("MIN-1-21 Project")
        window.iconbitmap("image.ico")
        window.geometry('305x100+550+200')
        subscribeToNewTariffWindow(window,idTariff,details)

    lbl = Label(window, text="- Menu -\n"
                            "Please, select the menu option to work with the program\n", background="#58D69B")
    lbl.grid(column=0, row=1, pady=(10, 0))
    btn = Button(window, text="My Tariff", command=lambda: clicked1(col))
    btn.grid(column=0, row=2, pady=(10, 0))
    btn = Button(window, text="My balance", command=lambda: clicked2(col))
    btn.grid(column=0, row=3)
    btn = Button(window, text="Subscribe to Tariff", command=lambda: clicked3(col))
    btn.grid(column=0, row=4)
    btn = Button(window, text="Exit the program", command=lambda: exit())
    btn.grid(column=0, row=5)
    
def emps_menu(col):
    workbook = load_workbook(filename="Users.xlsx")
    EmployeesSheet = workbook["Employees"]
    ClientsSheet = workbook["Clients"]
    VVX = load_workbook(filename="Tariffs.xlsx")
    Tariffs = VVX["Tariffs"]
    window = Tk()
    window.title("Employee's menu")
    window.geometry('400x380+550+200')
    window.resizable(width=False, height=False)
    window.configure(bg='#6791DC')
    lbl = Label(window, text="- Employee menu -\n"
                            "Please, select the menu option to work with the program\n", background='#6791DC')
    lbl.grid(column=0, row=1, pady=(10, 0),padx=50)

    def clicked01():
        window = Toplevel()
        window.title("MIN-1-21 Project")
        window.iconbitmap("image.ico")
        window.geometry('400x300+550+200')
        users_list(window,t='Client')

    def clicked02():
        window = Toplevel()
        Search=StringVar(window)
        window.title("MIN-1-21 Project")
        window.iconbitmap("image.ico")
        window.geometry('400x250+550+200')
        searchClientWindow(window,Search)

    def clicked03():
        window = Toplevel()
        Search=IntVar(window)
        window.title("MIN-1-21 Project")
        window.iconbitmap("image.ico")
        window.geometry('400x300+550+200')
        historyUserWindow(window,Search)

    def clicked04():
        window = Toplevel()
        window.title("MIN-1-21 Project")
        window.iconbitmap("image.ico")
        window.geometry('400x300+550+200')
        showTariffs(window)

    def clicked05():
        window = Toplevel()
        window.title("MIN-1-21 Project")
        window.iconbitmap("image.ico")
        window.geometry('400x300+550+200')
        viewListOfReqest(window)

    def clicked06():
        window = Toplevel()
        window.title("MIN-1-21 Project")
        window.iconbitmap("image.ico")
        window.geometry('400x300+550+200')
        sortClients(window)

    def clicked07():
        window = Toplevel()
        window.title("MIN-1-21 Project")
        window.iconbitmap("image.ico")
        window.geometry('400x300+550+200')
        stats(window)

    def clicked08():
        window = Toplevel()
        Search=IntVar(window)
        Amount=IntVar(window)
        window.title("MIN-1-21 Project")
        window.iconbitmap("image.ico")
        window.geometry('400x300+550+200')
        addBalanceWindow(window,Search,Amount)

    btn = Button(window, text='List of clients', command=lambda: clicked01())
    btn.grid(column=0, row=2, pady=(10, 0),padx=50)
    btn = Button(window, text="Search ", command=lambda: clicked02())
    btn.grid(column=0, row=3)
    btn = Button(window, text="Customer history", command=lambda: clicked03())
    btn.grid(column=0, row=4)
    btn = Button(window, text="Tariffs", command=lambda: clicked04())
    btn.grid(column=0, row=5)
    btn = Button(window, text="Issuing tariffs", command=lambda: clicked05())
    btn.grid(column=0, row=6)
    btn = Button(window, text="Sort clients", command=lambda: clicked06())
    btn.grid(column=0, row=7)
    btn = Button(window, text="Statistics of clients", command=lambda: clicked07())
    btn.grid(column=0, row=8)
    btn = Button(window, text="Add money to Client's balance", command=lambda: clicked08())
    btn.grid(column=0, row=9)
    btn = Button(window, text="Exit the program", command=lambda: exit())
    btn.grid(column=0, row=10)

def dirs_menu(col):
    window = Tk()
    window.title("Director's menu")
    window.geometry('305x250+550+200')
    window.resizable(width=False, height=False)
    window.configure(bg='#58D69B')
    lbl = Label(window, text="- Director menu -\n"
                            "Please, select the menu option to work with the program\n", background="#FF7860")
    lbl.grid(column=0, row=1, pady=(10, 0))

    def clicked001():
        window = Toplevel()
        window.title("MIN-1-21 Project")
        window.iconbitmap("image.ico")
        window.geometry('305x250+550+200')
        users_list(window,t='Employee')

    def clicked002():
        window1 = Toplevel()
        Login=StringVar(window1)
        Password=StringVar(window1)
        Name=StringVar(window1)
        Position=StringVar(window1)
        window1.title("MIN-1-21 Project")
        window1.iconbitmap("image.ico")
        window1.geometry('305x250+550+200')
        createEmployeeWindow(window1,Login,Password,Name,Position)

    def clicked004():
        window = Toplevel()
        Salary=IntVar(window)
        id=IntVar(window)
        window.title("MIN-1-21 Project")
        window.iconbitmap("image.ico")
        window.geometry('305x250+550+200')
        changeSalaryWindow(window,Salary,id)

    btn = Button(window, text="Worker list", command=lambda: clicked001())
    btn.grid(column=0, row=2, pady=(20, 0),padx=50)
    btn = Button(window, text="Add worker", command=lambda: clicked002())
    btn.grid(column=0, row=3)
    btn = Button(window, text="Salary", command=lambda: clicked004())
    btn.grid(column=0, row=4)
    btn = Button(window, text="Exit the program", command=lambda: exit(window))
    btn.grid(column=0, row=5)    
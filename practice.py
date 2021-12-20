from tkinter import *
from typing import *
import csv as v
from openpyxl import *
from colorama import init, Fore, Back, Style
import pandas as pd
root = Tk()
root.title("MIN-1-21 Project")
root.iconbitmap("image.ico")

workbook=load_workbook(filename="Users.xlsx")
EmployeesSheet=workbook["Employees"]
ClientsSheet=workbook["Clients"]
VVX=load_workbook(filename="Tariffs.xlsx")
Tariffs=VVX["Tariffs"]

Login=StringVar()
Password=StringVar()

def check(type,top,top2):
    auth1=Login.get()
    auth2=Password.get()
    CHECK1=False
    CHECK2=False
    for col in type.iter_rows(min_row=2,values_only=True):
        if col[2]==auth1:
            CHECK1=True
        if col[3]==auth2:
            CHECK2=True
    if (CHECK1==True) and (CHECK2==True):
        print("Success")
        top.destroy()
        top2.destroy()
    else:
        print("Failure")

def account_auth(type,ttt):
    enterwindow = Toplevel()
    enterwindow.title("MIN-1-21 Project")
    enterwindow.iconbitmap("image.ico")
    enterwindow.geometry("300x300")
    LoginField=Entry(enterwindow, textvariable=Login).pack()
    PasswordField=Entry(enterwindow, textvariable=Password).pack()
    Enter_Button = Button(enterwindow, text="Enter",command=lambda t=type,tt=enterwindow,t4=ttt: check(t,tt,t4)).pack()

def type_choice():
    authwindow = Toplevel()
    authwindow.geometry("300x300")
    authwindow.title("MIN-1-21 Project")
    authwindow.iconbitmap("image.ico")
    Emp_Button = Button(authwindow, text="Employee", command=lambda ttt=authwindow, m=EmployeesSheet: account_auth(m,ttt))
    Emp_Button.pack()
    Cli_Button = Button(authwindow, text="Client", command=lambda ttt=authwindow, m=ClientsSheet: account_auth(m,ttt))
    Cli_Button.pack()

def reg_client():
    enterwindow = Toplevel()
    enterwindow.title("MIN-1-21 Project")
    enterwindow.iconbitmap("image.ico")
    enterwindow.geometry("300x300")
    LoginField=Entry(enterwindow, textvariable=Login).pack()
    PasswordField=Entry(enterwindow, textvariable=Password).pack()
    
# def reg_c

def Reg_Auth():
    RegAuth=Toplevel()
    RegAuth.geometry("300x300")
    RegAuth.title("MIN-1-21 Project")
    RegAuth.iconbitmap("image.ico")
    Reg_Button = Button(RegAuth, text="Employee", command=lambda a=RegAuth: reg_client())
    Reg_Button.pack()
    Auth_Button = Button(RegAuth, text="Client", command=lambda a=RegAuth: type_choice())
    Auth_Button.pack()







root.mainloop()

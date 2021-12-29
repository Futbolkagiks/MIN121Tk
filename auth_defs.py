from tkinter import *
import tkinter
from typing import *
import csv as v
from matplotlib.pyplot import text
from openpyxl import *
from colorama import init, Fore, Back, Style
import pandas as pd
from tkinter import messagebox
from menus import *
# from menu_functions import *

workbook=load_workbook(filename="Users.xlsx")
EmployeesSheet=workbook["Employees"]
ClientsSheet=workbook["Clients"]
VVX=load_workbook(filename="Tariffs.xlsx")
Tariffs=VVX["Tariffs"]
root = Tk()
root.withdraw()

Name=StringVar()
Login=StringVar()
Password=StringVar()

def create_user(window,n,l,p):
    workbook=load_workbook(filename="Users.xlsx")
    ClientsSheet=workbook["Clients"]
    ac1=n.get()
    ac2=l.get()
    ac3=p.get()
    print(ac3,ac1,ac2)
    new_account=[int(len(ClientsSheet["A"])),ac1,ac2,ac3]
    if (ac1=='') or (ac2=='') or (ac3==''):
        messagebox.showinfo("Error", "One of the fields is not filled in")
    else:
        if login_check_2(ClientsSheet,window)==False:
            ClientsSheet.append(new_account)
            n.set('')
            l.set('')
            p.set('')
        workbook.save("Users.xlsx")
    
def login_check_2(type,window):
    LoginCheck=False
    while True:
        auth1=Login.get()
        for col in type.iter_rows(min_row=2,values_only=True):
            if auth1==col[2]:
                LoginCheck=True
        if LoginCheck==True:
            messagebox.showinfo("Error", "This Login is already in use!")
        else:
            messagebox.showinfo("Success","New account has been created")
            window.destroy()
        break
    return LoginCheck

def check(t,top,top2,a):
    workbook=load_workbook(filename="Users.xlsx")
    EmployeesSheet=workbook["Employees"]
    ClientsSheet=workbook["Clients"]
    auth1=Login.get()
    auth2=Password.get()
    CHECK1=False
    CHECK2=False
    if t=="Client":
        account_type=ClientsSheet
    elif t=="Employee":
        account_type=EmployeesSheet
    for col in account_type.iter_rows(min_row=2,values_only=True):
        if col[2]==auth1:
            CHECK1=True
        if col[3]==auth2:
            CHECK2=True
            global details
            details=col
    if (CHECK1==True) and (CHECK2==True):
        print("Success")
        top.destroy()
        top2.destroy()
        a.destroy()
        if t=="Client":
            users_menu(details)
        elif t=="Employee":
            if auth1=="L123" and auth2=="P123":
                dirs_menu(details)
            else:
                emps_menu(details)
    else:
        print("Failure")

def account_auth(type,ttt,a):
    enterwindow = Toplevel()
    enterwindow.title("MIN-1-21 Project")
    enterwindow.iconbitmap("image.ico")
    enterwindow.geometry('300x100+550+200')
    AuthLabel=Label(enterwindow,text="Authentication Window").grid(columnspan=2,row=0)
    LoginField=Entry(enterwindow, textvariable=Login).grid(column=1,row=1)
    LoginLabel=Label(enterwindow,text="Login").grid(column=0,row=1,padx=50)
    PasswordField=Entry(enterwindow, textvariable=Password, show='*').grid(column=1,row=2)
    PasswordLabel=Label(enterwindow,text="Password").grid(column=0,row=2)
    Enter_Button = Button(enterwindow, text="Enter",command=lambda t=type,tt=enterwindow,t4=ttt,aa=a: check(t,tt,t4,aa)).grid(columnspan=2,row=3)

def type_choice(a):
    authwindow = Toplevel()
    authwindow.geometry('300x100+550+200')
    authwindow.title("MIN-1-21 Project")
    authwindow.iconbitmap("image.ico")
    authwindow.configure(bg='#16C2A1')
    Emp_Button = Button(authwindow, text="Employee", command=lambda ttt=authwindow, m="Employee",aa=a: account_auth(m,ttt,aa))
    Emp_Button.grid(column=0,row=0,pady=40,padx=60)
    Cli_Button = Button(authwindow, text="Client", command=lambda ttt=authwindow, m="Client",aa=a: account_auth(m,ttt,aa))
    Cli_Button.grid(column=1,row=0)

def reg_client(a):
    enterwindow = Toplevel()
    enterwindow.title("MIN-1-21 Project")
    enterwindow.iconbitmap("image.ico")
    enterwindow.geometry('300x120+550+200')
    Name1=StringVar(enterwindow)
    Login2=StringVar(enterwindow)
    Password3=StringVar(enterwindow)
    RegistrationLabel=Label(enterwindow, text="Registration window").grid(columnspan=2,row=0)
    NameField=Entry(enterwindow, textvariable=Name1).grid(column=1,row=1)
    NameLabel=Label(enterwindow,text="Name").grid(column=0,row=1,padx=50)
    LoginField=Entry(enterwindow, textvariable=Login2).grid(column=1,row=2)
    LoginLabel=Label(enterwindow,text="Login").grid(column=0,row=2)
    PasswordField=Entry(enterwindow, textvariable=Password3).grid(column=1,row=3)
    PasswordLabel=Label(enterwindow,text="Password").grid(column=0,row=3)
    Enter_Button = Button(enterwindow, text="Enter",command=lambda w=enterwindow,n=Name1,l=Login2,p=Password3: create_user(w,n,l,p)).grid(columnspan=2,row=4)

def Reg_Auth(root):
    RegAuthwindow=Toplevel()
    RegAuthwindow.title("MIN-1-21 Project")
    RegAuthwindow.iconbitmap("image.ico")
    RegAuthwindow.geometry('300x200+550+200')
    RegAuthwindow.configure(background='#58D432')
    Reg_Button = Button(RegAuthwindow, text="Registration", command=lambda a=RegAuthwindow: reg_client(a))
    Reg_Button.grid(column=0,row=0,pady=70,padx=50)
    Auth_Button = Button(RegAuthwindow, text="Authentication", command=lambda a=RegAuthwindow: type_choice(a))
    Auth_Button.grid(column=1,row=0)

